from openpyxl import load_workbook
import pandas as pd

def update_results(sheet,row,col,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    
        # Feed results to Table 1 & Fig1
        raw_data = pd.DataFrame(writer.sheets[sheet].values)
        results = raw_data.iloc[1:].set_index([0,1])
        if sheet != 'FigS1':
            results.index = pd.read_excel(path, sheet,index_col=[0,1]).index
        results.columns = raw_data.iloc[0,2:]
        results.loc[row,col] = values
        results.to_excel(writer,sheet)

def update_fig1(row,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    
        # Feed results to Table 1 & Fig1
        table1 = pd.read_excel(path, 'Table1 & Fig1',index_col=[0,1])
        table1.loc[row,['Biomass [Gt C]','Uncertainty']] = values
        table1.to_excel(writer,'Table1 & Fig1')

def update_figs2s3(row,col,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

        # Feed results to Fig. S2-S3
        raw_data = pd.DataFrame(writer.sheets['FigS2-S3'].values)
        results = raw_data.iloc[1:].set_index(0)
        results.index = pd.read_excel(path, 'FigS2-S3',index_col=0).index
        results.columns = raw_data.iloc[0,1:]
        results.loc[row,col] = values
        results.to_excel(writer,'FigS2-S3')


def update_MS_data(row,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        raw_data = pd.DataFrame(writer.sheets['Data mentioned in MS'].values)
        data = raw_data.iloc[1:].set_index(0)
        data.columns = raw_data.iloc[0,1:]
        data.loc[row,'Original Value'] = values
        data.to_excel(writer,'Data mentioned in MS')


def update_tableS1(row,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # Feed results to Table S1
        tableS1 = pd.read_excel(path, 'Table S1',index_col=[0,1])
        tableS1.loc[row,'Number of individuals'] = values
        tableS1.to_excel(writer,'Table S1')

def update_fig2a(row,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        # Feed results to Fig. 2A
        fig2a = pd.read_excel(path, 'Fig2A',index_col=[0,1])
        fig2a.loc[row,['Biomass [Gt C]','Uncertainty']] = values
        fig2a.to_excel(writer,'Fig2A')
        
def update_fig2c(row,col,values,path,round=False):
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = load_workbook(path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        
        # Feed results to Fig. 2B
        fig2b = pd.DataFrame(writer.sheets['Fig2C'].values)
        fig2b.ix[row,col] = values
        fig2b.to_excel(writer,sheet_name='Fig2C',index=False,header=False)